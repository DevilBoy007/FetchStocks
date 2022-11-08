import easygui
import time
import datetime as dt
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

'''CHROMEDRIVER SETUP'''

c_options=Options() # create options object for chrome
c_options.add_argument("--headless")   # to scrape web without displaying window

s = Service(ChromeDriverManager().install())    # set up manager service for chromedriver
driver = webdriver.Chrome(service = s, options = c_options) # instantiate chromedriver

'''EXCEL SETUP'''
PATH = './stocks.xlsx'  # set path to local local workbook
wbk = openpyxl.load_workbook(PATH)  # create workbook object
sheet = wbk.worksheets[0]   # specify the first worksheet

'''core functionality one'''

'''BEGIN PROGRAM'''

size = int(easygui.enterbox('enter # of workbook entries\n'))
print('writing data to workbook...\n')
for spaces in range(45):
    print('-', end='-')
for row in range(2,size+2):
    tickr = sheet.cell(row=row, column = 1).value
    print('searching for {}'.format(tickr))
    driver.get('https://google.com/search?q={}+stock'.format(tickr))
    time.sleep(1)
    try:
        element = driver.find_element(By.XPATH, '//*[@id="knowledge-finance-wholepage__entity-summary"]/div/g-card-section/div/g-card-section/div[2]/div[1]/span[1]/span/span[1]')
        currentPrice = element.text
        sheet.cell(row=row, column=2).value = currentPrice
        sheet.cell(row=row, column=3).value = (dt.datetime.now()).strftime('%c')
    except Exception as e:
        print('error occurred')
wbk.save(PATH)

'''core functionality one'''

'''BEGIN PROGRAM'''
sheet = wbk.worksheets[1]   # specify the secoond worksheet
row = 2 # input range begins at (2, 1)

while(True):
    tickr = easygui.enterbox("enter ticker symbol\nor type 'quit' to end: ")  # get ticker from user
    if(tickr=='quit'):
        quit()
    driver.get('https://google.com/search?q={}+stock'.format(tickr))  # query google
    time.sleep(1)
    try:
        print('searching markets...')
        currentPrice = driver.find_element(By.XPATH, '//*[@id="knowledge-finance-wholepage__entity-summary"]/div/g-card-section/div/g-card-section/div[2]/div[1]/span[1]/span/span[1]')
        currentPrice = currentPrice.text
        easygui.msgbox('{} is currently priced at: ${}'.format(tickr.upper(), currentPrice), title=tickr.upper())
        if(float(currentPrice.replace(',',''))>500):
            buy = 'yes'
        else:
            buy = 'no'
        sheet.cell(row=row, column=1).value = tickr.upper()
        sheet.cell(row=row, column=2).value = '${}'.format(currentPrice)
        sheet.cell(row=row, column=3).value = (dt.datetime.now()).strftime('%c')
        sheet.cell(row=row, column=4).value = buy
        wbk.save(PATH)
        row += 1
    except Exception as e:
        easygui.msgbox('Error occurred:\n{}'.format(e), title='ERROR!')

'''PROGRAM END'''
wbk.close()
driver.quit()

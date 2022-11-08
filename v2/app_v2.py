import easygui
import time
import datetime as dt
import openpyxl
import requests

'''EXCEL SETUP'''
ORIGINALPATH = './test.xlsx'  # set path to local local workbook
EXPORTPATH = './outputs/{}.xlsx'
wbk = openpyxl.load_workbook(ORIGINALPATH)  # create workbook object
col = 3 # all input happens in third column
sheetNumber = 0
'''BEGIN PROGRAM'''

while(True):
    sheet = wbk.worksheets[sheetNumber]   # specify the worksheet
    tickr = easygui.enterbox('enter ticker\nor type \'quit\' to end program')  # get ticker from user
    if(tickr.lower()=='quit'):
        quit()
    try:
        print('getting info for {}...\n'.format(tickr.upper()))

        ### call YHFinance API ###
        url = 'https://yh-finance.p.rapidapi.com/stock/v3/get-statistics'
        querystring = {'symbol': tickr.upper()}
        headers = {'x-rapidapi-host': 'yh-finance.p.rapidapi.com','x-rapidapi-key': '8d36fc9dacmsh905d9e293400d5bp1c6bedjsnb29677382b1b'}
        response = requests.request('GET', url, headers=headers, params=querystring)
        time = response.headers.get('date')
        response = response.json() # change response format to a dict we can work with

        ### create variables for target data ###
        name = (response.get('quoteType')).get('shortName') # C1
        ###   'tickr' is C2   ###
        ###   'time' is C3   ###
        price = ((response.get('price')).get('regularMarketPrice')).get('raw')    # C5
        if(((response.get('defaultKeyStatistics')).get('pegRatio'))):
            pegRatio = ((response.get('defaultKeyStatistics')).get('pegRatio')).get('raw')/100  # C6
        else:
            # set PEG to 0 to avoid errors in Excel document
            easygui.msgbox('No PEG ratio returned!\n *** 0 value will be used ***')
            pegRatio = 0
        forwardPE = ((response.get('defaultKeyStatistics')).get('forwardPE')).get('raw')  # C7
        epsYear0 = ((response.get('defaultKeyStatistics')).get('trailingEps')).get('raw')   # C8
        dps = ((response.get('financialData')).get('totalCashPerShare')).get('raw')  # C9
        ###   END variables   ###

    #   easygui.msgbox('{} is currently priced at: ${}'.format(tickr.upper(), price), title=tickr.upper())
        sheet.title = name
        sheet.cell(row=1, column=2).value = name    # merged cells must be referenced by top left cell, all others are read-only
        sheet.cell(row=2, column=col).value = tickr.upper()
        sheet.cell(row = 3, column=col).value = time # the time we made the request
        sheet.cell(row=3, column=4).value = (dt.datetime.now()).strftime('%c') # current system time
        sheet.cell(row=5, column=col).value = price
        sheet.cell(row=6, column=col).value = pegRatio
        sheet.cell(row=7, column=col).value = forwardPE
        sheet.cell(row=8, column=col).value = epsYear0
        sheet.cell(row=9, column=col).value = dps
        wbk.copy_worksheet(wbk.worksheets[sheetNumber])
        sheetNumber += 1
    except Exception as e:
        easygui.msgbox('Error occurred:\n{}'.format(e), title='ERROR!')
    wbk.save(EXPORTPATH.format((dt.datetime.now()).strftime('%x').replace('/', '-')))
'''PROGRAM END'''
wbk.close()
